#%%

import pandas as pd
import networkx as nx
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Paths
nodes_csv = r"C:\\Thesis code\manual_nodes_updated.csv"
edges_csv = r"C:\\Thesis code\edges_updated.csv"
excel_file = r"C:\\Documenten\big2.xlsx"
output_node_values = r"C:\\Thesis code\node_values_voltage.csv"

# --- Step 1: Read nodes and edges ---
print("Reading nodes CSV...")
nodes_df = pd.read_csv(nodes_csv)
nodes = nodes_df.iloc[:, 0].astype(str).tolist()
print(f"Number of nodes: {len(nodes)}")

print("Reading edges CSV...")
edges_df = pd.read_csv(edges_csv)
edges = [(str(row[0]), str(row[1])) for _, row in edges_df.iloc[:, :2].iterrows()]
print(f"Number of edges: {len(edges)}")

#%%

# --- Step 2: Identify first relevant column per node ---
print("Reading first 2 rows of Excel to identify relevant columns...")
header_df = pd.read_excel(excel_file, sheet_name=0, nrows=2, header=None)

# Map node number -> first relevant column index
node_to_column = {}
for col in header_df.columns:
    if str(header_df.iloc[1, col]) == 'u, Magnitude in p.u.':
        header = str(header_df.iloc[0, col])
        node_number = header.split('\\')[0]  # take only the first number
        if node_number not in node_to_column:
            node_to_column[node_number] = col
            print(f"Selected column {col} for node number {node_number}, header: {header}")

print(f"Total selected node columns: {len(node_to_column)}")

#%%
keep_cols = {71, 73, 76, 79, 82, 85, 90, 103, 106, 114, 117, 144, 158, 170, 173}

# Remove columns 50–175 except the ones in keep_cols
for node, col in list(node_to_column.items()):
    if 50 <= col <= 175 and col not in keep_cols:
        del node_to_column[node]

print(f"Total selected node columns after deletion: {len(node_to_column)}")
#%%
# Extract node numbers from nodes CSV (first number before '_')
csv_node_numbers = set(node.split('_')[0] for node in nodes)

# Columns whose node numbers are not in the CSV nodes
non_matching_nodes = [node_number for node_number in node_to_column.keys() if node_number not in csv_node_numbers]

print("Columns with node numbers that do not match any node in CSV:")
print(non_matching_nodes)

#%%

# --- Step 3: Read only selected columns efficiently using openpyxl ---
print("Reading only selected columns from Excel using openpyxl (read-only mode)...")
wb = load_workbook(excel_file, read_only=True)
ws = wb.active

# Map column indices (0-based) to Excel letters
selected_cols = list(node_to_column.values())
col_idx_to_letter = {col: get_column_letter(col + 1) for col in selected_cols}

# Initialize data dictionary
data_dict = {col_idx: [] for col_idx in selected_cols}

for row_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True), 1):
    for col_idx in selected_cols:
        data_dict[col_idx].append(row[col_idx])
    if row_idx % 500 == 0:
        print(f"Read {row_idx} rows...")

# Convert to DataFrame
data_df = pd.DataFrame(data_dict)

# Map column index -> DataFrame column name (just using same indices)
col_index_to_name = {col: col for col in selected_cols}
print(f"Data shape of selected columns: {data_df.shape}")

#%%

# --- Step 4: Compute node values ---
print("Computing node values...")
node_values = {}
for i, node in enumerate(nodes, 1):
    node_number = node.split('_')[0]
    if node_number in node_to_column:
        col_idx = node_to_column[node_number]
        values = pd.to_numeric(data_df[col_idx], errors='coerce')
        values = values[values > 1.05] - 1.05
        node_values[node] = values.sum()
    else:
        node_values[node] = None

    if i % 50 == 0 or i == len(nodes):
        print(f"Processed {i}/{len(nodes)} nodes...")

#%%

# --- Step 5: Create graph ---
G = nx.Graph()
for node in nodes:
    G.add_node(node, value=node_values[node])
G.add_edges_from(edges)


#%%
import matplotlib.colors as mcolors
# --- Step 6: Draw graph ---
values = [node_values[node] if node_values[node] is not None else 0 for node in G.nodes()]
manual_max = 70  # <-- your chosen upper bound
vmin = 0         # or another lower bound you prefer
vmax = manual_max  # instead of max(values)

norm = mcolors.Normalize(vmin=vmin, vmax=vmax)
fig, ax = plt.subplots(figsize=(16, 10))   # Larger figure for more space
pos = nx.kamada_kawai_layout(G)
for key in pos:
    pos[key] = pos[key] * 2.5 

# Draw nodes, edges, and labels
nodes = nx.draw_networkx_nodes(
    G, pos,
    node_color=values,
    cmap=plt.cm.Reds,
    node_size=80,     # smaller nodes
    ax=ax
)
nx.draw_networkx_nodes(G, pos, node_color='none', edgecolors='black', node_size=81, linewidths=0.1, ax=ax)
nx.draw_networkx_edges(G, pos, alpha=0.4, ax=ax)
#nx.draw_networkx_labels(G, pos, font_size=6, font_color="black", ax=ax)  # smaller labels

# Add colorbar (smaller scale)
#sm = plt.cm.ScalarMappable(cmap=plt.cm.Reds, norm=plt.Normalize(vmin=min(values), vmax=max(values)))
sm = plt.cm.ScalarMappable(cmap=plt.cm.Reds, norm=norm)
sm.set_array([])
cbar = plt.colorbar(sm, ax=ax, fraction=0.03, pad=0.04)  # makes it smaller and closer
cbar.set_label("Summed Voltage violations", fontsize=10)
cbar.ax.tick_params(labelsize=8)  # smaller tick labels

ax.set_title("Network Graph with Node Values", fontsize=14, pad=20)
ax.axis("off")

plt.tight_layout()
plt.show()

#%%

# --- Step 7: Save node values ---
node_values_df = pd.DataFrame(list(node_values.items()), columns=['Node', 'Value'])
node_values_df.to_csv(output_node_values, index=False)
print(f"Node values saved to {output_node_values}")
# %%
